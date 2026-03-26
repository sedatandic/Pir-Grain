from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from database import db
from auth import get_current_user
from datetime import datetime, timezone
import openpyxl
import io
import tempfile
import os

router = APIRouter(prefix="/api/port-lineups", tags=["port-lineups"])

HEADER_MARKER = "GEMİ ADI"


def parse_port_report(file_bytes: bytes) -> list:
    """Parse the daily port report Excel file and return structured data."""
    # Save to temp file (openpyxl needs file path or file-like object)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
        all_reports = []

        for sheet_name in wb.sheetnames:
            report_date = sheet_name.strip()
            ws = wb[sheet_name]

            ports_data = []
            current_port = None
            current_vessels = []
            expecting_port_name = False

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
                # Get cell values (columns B through J = indices 1-9)
                cells = {}
                for cell in row:
                    if cell.value is not None and cell.column >= 2 and cell.column <= 10:
                        cells[cell.column] = cell.value

                if not cells:
                    continue

                # Check if this is a header row
                col_b_val = cells.get(2, "")
                if isinstance(col_b_val, str) and col_b_val.strip() == HEADER_MARKER:
                    # Save previous port section
                    if current_port and current_vessels:
                        ports_data.append({
                            "portName": current_port,
                            "vessels": current_vessels
                        })
                    current_port = None
                    current_vessels = []
                    expecting_port_name = True
                    continue

                # If we're expecting a port name (row after header)
                if expecting_port_name:
                    if col_b_val and isinstance(col_b_val, str) and col_b_val.strip():
                        # Check it's not a vessel row (vessel rows have data in col C too)
                        col_c_val = cells.get(3)
                        if not col_c_val:
                            current_port = col_b_val.strip()
                            expecting_port_name = False
                            continue
                        else:
                            # It's actually a vessel row, port name might be missing
                            # Use a default
                            current_port = col_b_val.strip()
                            expecting_port_name = False
                            # Fall through to process as vessel

                # Process vessel data row
                if current_port is not None:
                    vessel_name = cells.get(2, "")
                    loading_port = cells.get(3, "")
                    arrival_date_raw = cells.get(4)
                    status = cells.get(5, "")
                    operation = cells.get(6, "")
                    cargo = cells.get(7, "")
                    bl_tonnage = cells.get(8)
                    buyer = cells.get(9, "")
                    seller = cells.get(10, "")

                    # Must have at least vessel name or loading port to be valid
                    if not vessel_name and not loading_port:
                        continue

                    # Parse arrival date
                    arrival_date_str = ""
                    if isinstance(arrival_date_raw, datetime):
                        arrival_date_str = arrival_date_raw.strftime("%d.%m.%Y")
                    elif isinstance(arrival_date_raw, str) and arrival_date_raw.strip():
                        arrival_date_str = arrival_date_raw.strip()

                    # Parse tonnage
                    tonnage = None
                    if bl_tonnage is not None:
                        try:
                            tonnage = float(bl_tonnage)
                        except (ValueError, TypeError):
                            tonnage = None

                    vessel = {
                        "vesselName": str(vessel_name).strip() if vessel_name else "",
                        "loadingPort": str(loading_port).strip() if loading_port else "",
                        "arrivalDate": arrival_date_str,
                        "status": str(status).strip() if status else "",
                        "operation": str(operation).strip() if operation else "",
                        "cargo": str(cargo).strip() if cargo else "",
                        "blTonnage": tonnage,
                        "buyer": str(buyer).strip() if buyer else "",
                        "seller": str(seller).strip() if seller else "",
                    }
                    current_vessels.append(vessel)

            # Save last port section
            if current_port and current_vessels:
                ports_data.append({
                    "portName": current_port,
                    "vessels": current_vessels
                })

            if ports_data:
                all_reports.append({
                    "reportDate": report_date,
                    "ports": ports_data
                })

        wb.close()
        return all_reports
    finally:
        os.unlink(tmp_path)


@router.post("/upload")
async def upload_port_report(file: UploadFile = File(...), current_user=Depends(get_current_user)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")

    file_bytes = await file.read()
    try:
        reports = parse_port_report(file_bytes)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")

    if not reports:
        raise HTTPException(status_code=400, detail="No valid port data found in the file")

    # Clear existing data and insert new
    collection = db["port_lineups"]
    collection.delete_many({})

    now = datetime.now(timezone.utc).isoformat()
    for report in reports:
        report["uploadedAt"] = now
        report["uploadedBy"] = current_user.get("username", "unknown")

    collection.insert_many(reports)

    return {
        "message": f"Successfully uploaded {len(reports)} daily reports",
        "dates": [r["reportDate"] for r in reports],
        "totalPorts": sum(len(r["ports"]) for r in reports),
        "totalVessels": sum(
            sum(len(p["vessels"]) for p in r["ports"])
            for r in reports
        )
    }


@router.get("/dates")
async def get_report_dates(current_user=Depends(get_current_user)):
    collection = db["port_lineups"]
    dates = collection.distinct("reportDate")

    # Sort dates descending (newest first)
    def parse_date(d):
        try:
            return datetime.strptime(d.strip(), "%d.%m.%Y")
        except ValueError:
            return datetime.min

    dates.sort(key=parse_date, reverse=True)
    return {"dates": dates}


@router.get("/report/{report_date}")
async def get_report(report_date: str, current_user=Depends(get_current_user)):
    collection = db["port_lineups"]
    doc = collection.find_one({"reportDate": report_date}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail=f"No report found for date {report_date}")
    return doc


@router.get("/summary")
async def get_summary(current_user=Depends(get_current_user)):
    """Get a summary of latest report: port names, vessel counts."""
    collection = db["port_lineups"]

    # Get the latest report date
    dates = collection.distinct("reportDate")
    if not dates:
        return {"latestDate": None, "ports": [], "totalVessels": 0}

    def parse_date(d):
        try:
            return datetime.strptime(d.strip(), "%d.%m.%Y")
        except ValueError:
            return datetime.min

    dates.sort(key=parse_date, reverse=True)
    latest_date = dates[0]

    doc = collection.find_one({"reportDate": latest_date}, {"_id": 0})
    if not doc:
        return {"latestDate": latest_date, "ports": [], "totalVessels": 0}

    port_summary = []
    total_vessels = 0
    for port in doc.get("ports", []):
        vessel_count = len(port.get("vessels", []))
        total_vessels += vessel_count
        port_summary.append({
            "portName": port["portName"],
            "vesselCount": vessel_count
        })

    return {
        "latestDate": latest_date,
        "ports": port_summary,
        "totalVessels": total_vessels,
        "totalDates": len(dates)
    }


# ─── Monthly Line-Up ───

UPLOAD_DIR = "/app/backend/uploads/monthly_lineups"
os.makedirs(UPLOAD_DIR, exist_ok=True)

monthly_col = db["monthly_lineups"]


def parse_monthly_excel(file_bytes: bytes, filename: str):
    """Parse monthly lineup Excel and return all sheets as structured tables."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                vals = [str(c).strip() if c is not None else "" for c in row]
                if not any(vals):
                    continue
                if not headers:
                    headers = vals
                else:
                    rows.append(dict(zip(headers, vals)))
            if headers and rows:
                sheets.append({"sheetName": sheet_name, "headers": headers, "rows": rows})
        wb.close()
        return sheets
    finally:
        os.unlink(tmp_path)


@router.post("/monthly/upload")
async def upload_monthly_lineup(file: UploadFile = File(...), current_user=Depends(get_current_user)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files supported")
    file_bytes = await file.read()
    try:
        sheets = parse_monthly_excel(file_bytes, file.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse: {str(e)}")
    if not sheets:
        raise HTTPException(status_code=400, detail="No data found in Excel file")

    # Store file on disk
    from bson import ObjectId
    doc_id = ObjectId()
    ext = file.filename.rsplit('.', 1)[-1] if '.' in file.filename else 'xlsx'
    stored_name = f"monthly_{doc_id}.{ext}"
    with open(os.path.join(UPLOAD_DIR, stored_name), "wb") as f:
        f.write(file_bytes)

    doc = {
        "_id": doc_id,
        "fileName": file.filename,
        "storedFileName": stored_name,
        "sheets": sheets,
        "uploadedAt": datetime.now(timezone.utc).isoformat(),
        "uploadedBy": current_user.get("username", "unknown"),
    }
    monthly_col.insert_one(doc)
    return {
        "id": str(doc_id),
        "fileName": file.filename,
        "sheetCount": len(sheets),
        "totalRows": sum(len(s["rows"]) for s in sheets),
    }


@router.get("/monthly/list")
async def list_monthly_lineups(current_user=Depends(get_current_user)):
    docs = list(monthly_col.find({}, {"sheets": 0}).sort("uploadedAt", -1))
    for d in docs:
        d["id"] = str(d.pop("_id"))
    return docs


@router.get("/monthly/{doc_id}")
async def get_monthly_lineup(doc_id: str, current_user=Depends(get_current_user)):
    from bson import ObjectId
    doc = monthly_col.find_one({"_id": ObjectId(doc_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    doc["id"] = str(doc.pop("_id"))
    return doc


@router.delete("/monthly/{doc_id}")
async def delete_monthly_lineup(doc_id: str, current_user=Depends(get_current_user)):
    from bson import ObjectId
    doc = monthly_col.find_one({"_id": ObjectId(doc_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    stored = doc.get("storedFileName")
    if stored:
        fpath = os.path.join(UPLOAD_DIR, stored)
        if os.path.exists(fpath):
            os.remove(fpath)
    monthly_col.delete_one({"_id": ObjectId(doc_id)})
    return {"message": "Deleted"}


@router.get("/monthly/{doc_id}/download")
async def download_monthly_lineup(doc_id: str, current_user=Depends(get_current_user)):
    from bson import ObjectId
    from fastapi.responses import FileResponse
    doc = monthly_col.find_one({"_id": ObjectId(doc_id)})
    if not doc or not doc.get("storedFileName"):
        raise HTTPException(status_code=404, detail="File not found")
    fpath = os.path.join(UPLOAD_DIR, doc["storedFileName"])
    if not os.path.exists(fpath):
        raise HTTPException(status_code=404, detail="File not on disk")
    return FileResponse(fpath, filename=doc.get("fileName", "monthly_lineup.xlsx"), media_type="application/octet-stream")
